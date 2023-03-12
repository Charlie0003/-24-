from collections import Counter
from openpyxl import Workbook

wb = Workbook()
ws = wb.active


class Write:
    def __init__(self):
        # 询问路径并查找
        self.file_path = input('请输入路径，不输入则放E盘')
        self.file_name = input('请输入文件名，不输入则首字母')
        self.get_path()
        self.word_list = []
        self.get_message()

    def get_path(self):
        # 询问路径
        if self.file_path == '':
            self.file_path = 'E:\\'
        # 询问文件名
        if self.file_name == '':
            self.file_name = '首字母.xlsx'
        else:
            self.file_name = f'{self.file_name}.xlsx'
        # 连接起来得到路径
        self.else_path = f"{self.file_path}exam.txt"
        self.file_path = f"{self.file_path}{self.file_name}"

    def run(self):
        while True:
            # 输入新单词
            new_word = input('请输入新单词')
            if new_word.strip() == 'D':
                delete_things = input('删除哪个单词')
                self.word_list.remove(delete_things)
            elif new_word.strip() != 'Q':
                self.word_list.append(new_word)
                print(new_word)
                print(self.word_list)
            else:
                self.write_message()
                self.count()
                break

    def get_message(self):
        with open(self.else_path, 'r') as file:
            lines = file.readlines()
        for line in lines:
            self.word_list.append(line.strip())

    def write_message(self):
        with open(self.else_path, 'w') as file:
            for i in self.word_list:
                file.write(f'{i}\n')

    def count(self):
        counter = Counter(self.word_list)
        word_times = counter.most_common(99999)
        number = 1
        for key, value in word_times:
            ws[f'A{number}'] = key
            ws[f'B{number}'] = value
            number += 1
        wb.save(f'{self.file_path}')


if __name__ == '__main__':
    wr = Write()
    wr.run()